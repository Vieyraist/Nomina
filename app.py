from flask import Flask, render_template, request, redirect, send_file, session
from datetime import datetime
from openpyxl import Workbook
import psycopg2
import sqlite3
import os
from functools import wraps

app = Flask(__name__)

# 🔐 CONFIG SEGURIDAD
app.secret_key = os.getenv("SECRET_KEY", "dev_key")

# 🔐 USUARIO ÚNICO
USERNAME = "Admin"
PASSWORD = "mexico24"

# ================== CONFIG ==================
ENV = os.getenv("ENV", "dev")


# ================== LOGIN REQUIRED ==================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            return redirect("/login")
        return f(*args, **kwargs)
    return decorated_function
    
    
    # ===== VER NOMINA ===============
from openpyxl import load_workbook

@app.route("/ver_nomina/<nombre>")
@login_required
def ver_nomina(nombre):

    ruta = os.path.join("nominas", nombre)

    wb = load_workbook(ruta)
    ws = wb.active

    data = []

    for row in ws.iter_rows(values_only=True):
        # limpiar filas vacías
        if any(cell is not None for cell in row):
            data.append(list(row))

    # 🔍 buscar encabezado real
    header_index = 0
    for i, row in enumerate(data):
        if "Zona" in row:
            header_index = i
            break

    headers = data[header_index]

    # limpiar encabezados None
    headers = [h if h else "" for h in headers]

    filas = data[header_index + 1:]

    # eliminar filas completamente vacías
    filas = [f for f in filas if any(f)]

    return render_template(
        "ver_nomina.html",
        headers=headers,
        filas=filas,
        archivo=nombre
    )# ================== HOME (REDIRECCIÓN INTELIGENTE) ==================
@app.route("/")
def home():
    if "user" in session:
        return redirect("/index")
    return redirect("/login")

# ================== LOGIN ==================
@app.route("/login", methods=["GET", "POST"])
def login():

    # 🔁 si ya está logueado → no volver a login
    if "user" in session:
        return redirect("/")

    if request.method == "POST":
        user = request.form["username"]
        pwd = request.form["password"]

        if user == USERNAME and pwd == PASSWORD:
            session["user"] = user
            return redirect("/")

        return render_template("login.html", error=True)

    return render_template("login.html", error=False)

# ================== LOGOUT ==================
@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect("/login")

# ================== CONEXION DB ==================
def get_db():
    if ENV == "dev":
        conn = sqlite3.connect("database.db")
        conn.row_factory = sqlite3.Row
        return conn
    else:
        return psycopg2.connect(os.getenv("DATABASE_URL"))

# ================== INIT SQLITE ==================
def init_sqlite():
    if ENV == "dev":
        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
        CREATE TABLE IF NOT EXISTS empleados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            zona TEXT,
            puesto TEXT,
            nombre TEXT,
            salario REAL,
            tarifa REAL,
            fecha_ingreso TEXT
        )
        """)

        conn.commit()
        conn.close()

init_sqlite()

# ================== UTIL ==================
def to_float(valor):
    try:
        return float(valor)
    except:
        return 0


def calcular_antiguedad(fecha_str):
    if not fecha_str:
        return "-"

    try:
        fecha = datetime.strptime(fecha_str, "%Y-%m-%d")
        hoy = datetime.now()
        dias = (hoy - fecha).days

        años = dias // 365
        meses = (dias % 365) // 30

        return f"{años} años, {meses} meses"
    except:
        return "-"




# ================== OBTENER EMPLEADOS ==================
def obtener_empleados():
    conn = get_db()
    cur = conn.cursor()

    if ENV == "dev":
        cur.execute("""
        SELECT id, zona, puesto, nombre, salario, tarifa, fecha_ingreso
        FROM empleados
        ORDER BY puesto
        """)
    else:
        cur.execute("""
        SELECT id, zona, puesto, nombre, salario, tarifa, fecha_ingreso
        FROM empleados
        ORDER BY 
            CASE 
                WHEN LOWER(puesto) = 'residente' THEN 1
                WHEN LOWER(puesto) = 'cabo' THEN 2
                WHEN LOWER(puesto) = 'pintor' THEN 3
                WHEN LOWER(puesto) = 'soldador' THEN 4
                WHEN LOWER(puesto) = 'ayudante' THEN 5
                ELSE 6
            END
        """)

    data = cur.fetchall()

    cur.close()
    conn.close()

    empleados = []
    for e in data:
        empleados.append({
            "id": e[0],
            "zona": e[1],
            "puesto": e[2],
            "nombre": e[3],
            "salario": e[4],
            "tarifa": e[5],
            "fecha_ingreso": e[6],
            "antiguedad": calcular_antiguedad(e[6])
        })

    return empleados
    
# ================== INDEX ==================
@app.route("/index")
@login_required
def index():
    return render_template("index.html", empleados=obtener_empleados())

# ================== AGREGAR ==================
@app.route("/agregar_empleado", methods=["GET", "POST"])
@login_required
def agregar_empleado():
    if request.method == "POST":
        data = request.form

        conn = get_db()
        cur = conn.cursor()

        if ENV == "dev":
            cur.execute("""
                INSERT INTO empleados (zona, puesto, nombre, salario, tarifa, fecha_ingreso)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                data["zona"],
                data["puesto"],
                data["nombre"],
                float(data["salario"]),
                float(data.get("tarifa") or 0),
                data["fecha_ingreso"]
            ))
        else:
            cur.execute("""
                INSERT INTO empleados (zona, puesto, nombre, salario, tarifa, fecha_ingreso)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                data["zona"],
                data["puesto"],
                data["nombre"],
                float(data["salario"]),
                float(data.get("tarifa") or 0),
                data["fecha_ingreso"]
            ))

        conn.commit()
        cur.close()
        conn.close()

        return redirect("/")

    return render_template("agregar.html")

# ================== ELIMINAR ==================
@app.route("/eliminar/<int:id>")
@login_required
def eliminar(id):
    conn = get_db()
    cur = conn.cursor()

    if ENV == "dev":
        cur.execute("DELETE FROM empleados WHERE id = ?", (id,))
    else:
        cur.execute("DELETE FROM empleados WHERE id = %s", (id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect("/empleados")

# ================== EDITAR ==================
@app.route("/empleados", methods=["GET", "POST"])
@login_required
def empleados_view():

    if request.method == "POST":
        conn = get_db()
        cur = conn.cursor()

        empleados = obtener_empleados()

        for emp in empleados:
            zona = request.form.getlist(f"zona_{emp['id']}")[-1]
            puesto = request.form.getlist(f"puesto_{emp['id']}")[-1]
            nombre = request.form.getlist(f"nombre_{emp['id']}")[-1]
            salario = to_float(request.form.getlist(f"salario_{emp['id']}")[-1])
            tarifa = to_float(request.form.getlist(f"tarifa_{emp['id']}")[-1])
            fecha_ingreso = request.form.getlist(f"fecha_{emp['id']}")[-1]

            if ENV == "dev":
                cur.execute("""
                    UPDATE empleados
                    SET zona=?, puesto=?, nombre=?, salario=?, tarifa=?
                    WHERE id=?
                """, (zona, puesto, nombre, salario, tarifa, fecha_ingreso, emp["id"]))
            else:
                cur.execute("""
                    UPDATE empleados
                    SET zona=%s, puesto=%s, nombre=%s, salario=%s, tarifa=%s
                    WHERE id=%s
                """, (zona, puesto, nombre, salario, tarifa, fecha_ingreso, emp["id"]))

        conn.commit()
        cur.close()
        conn.close()

        return redirect("/empleados")

    return render_template("empleados.html", empleados=obtener_empleados())

# ================== HISTORIAL ==================
@app.route("/historial")
@login_required
def historial():
    carpeta = "nominas"
    if not os.path.exists(carpeta):
        os.makedirs(carpeta)

    archivos = os.listdir(carpeta)
    archivos.sort(reverse=True)

    return render_template("historial.html", archivos=archivos)

@app.route("/descargar/<nombre>")
@login_required
def descargar(nombre):
    ruta = os.path.join("nominas", nombre)
    return send_file(ruta, as_attachment=True)

@app.route("/eliminar_nomina/<nombre>")
@login_required
def eliminar_nomina(nombre):
    ruta = os.path.join("nominas", nombre)
    if os.path.exists(ruta):
        os.remove(ruta)
    return redirect("/historial")

# ================== GENERAR ==================
@app.route("/generar", methods=["POST"])
@login_required
def generar():
    try:
        empleados = obtener_empleados()

        if not os.path.exists("nominas"):
            os.makedirs("nominas")

        wb = Workbook()
        ws = wb.active

        headers = ["Zona","Puesto","Nombre","Base","Tarifa","Horas","Pago T.E","Otros","Descuentos","Total"]
        ws.append(headers)

        total_general = 0

        for emp in empleados:
            base = emp["salario"]
            tarifa = emp["tarifa"]

            horas = to_float(request.form.get(f"horas_{emp['id']}"))
            otros = to_float(request.form.get(f"otros_{emp['id']}"))
            desc = to_float(request.form.get(f"desc_{emp['id']}"))

            pago_te = horas * tarifa
            final = base + pago_te + otros - desc

            total_general += final

            ws.append([
                emp["zona"],
                emp["puesto"],
                emp["nombre"],
                base,
                tarifa,
                horas,
                pago_te,
                otros,
                desc,
                final
            ])

        ws.append(["", "", "", "", "", "", "", "", "TOTAL", total_general])

    # 🔥 AQUÍ VA TU CÓDIGO NUEVO (DENTRO DEL TRY)

        meses = {
            1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
            5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
            9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
        }

        hoy = datetime.now()
        fecha = f"{hoy.day}{meses[hoy.month]}{hoy.year}"

        archivo = f"nominas/Nomina_{fecha}.xlsx"

        wb.save(archivo)

        return send_file(archivo, as_attachment=True)

    except Exception as e:
        return f"❌ Error: {e}"
# ================== RUN ==================
if __name__ == "__main__":
    app.run(debug=True)