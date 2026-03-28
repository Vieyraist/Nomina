from flask import Flask, render_template, request, redirect, send_file
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import psycopg2
import os

app = Flask(__name__)

# ================== CONEXION DB ==================
def get_db():
    return psycopg2.connect(
        "postgresql://postgres.ahzihgurwtumohsmyfur:Omar1307jacky@aws-1-us-east-1.pooler.supabase.com:6543/postgres"
    )

# ================== UTIL ==================
def to_float(valor):
    try:
        return float(valor)
    except:
        return 0

# ================== OBTENER EMPLEADOS ==================
def obtener_empleados():
    conn = get_db()
    cur = conn.cursor()

    cur.execute("""
    SELECT id, zona, puesto, nombre, salario, tarifa 
    FROM empleados
    ORDER BY 
        CASE 
            WHEN LOWER(puesto) = 'residente' THEN 1
            WHEN LOWER(puesto) = 'cabo' THEN 2
            WHEN LOWER(puesto) = 'pintor' THEN 3
            WHEN LOWER(puesto) = 'soldador' THEN 4
            WHEN LOWER(puesto) = 'ayudante' THEN 5
            ELSE 6
        END
    """)

    data = cur.fetchall()

    cur.close()
    conn.close()

    empleados = []
    for e in data:
        empleados.append({
            "id": e[0],
            "zona": e[1],
            "puesto": e[2],
            "nombre": e[3],
            "salario": e[4],
            "tarifa": e[5],
        })

    return empleados

# ================== INDEX ==================
@app.route("/")
def index():
    return render_template("index.html", empleados=obtener_empleados())

# ================== AGREGAR ==================
@app.route("/agregar_empleado", methods=["GET", "POST"])
def agregar_empleado():
    if request.method == "POST":
        data = request.form

        conn = get_db()
        cur = conn.cursor()

        cur.execute("""
            INSERT INTO empleados (zona, puesto, nombre, salario, tarifa)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            data["zona"],
            data["puesto"],
            data["nombre"],
            float(data["salario"]),
            float(data.get("tarifa") or 0)
        ))

        conn.commit()
        cur.close()
        conn.close()

        return redirect("/")

    return render_template("agregar.html")

# ================== ELIMINAR ==================
@app.route("/eliminar/<int:id>")
def eliminar(id):
    conn = get_db()
    cur = conn.cursor()

    cur.execute("DELETE FROM empleados WHERE id = %s", (id,))

    conn.commit()
    cur.close()
    conn.close()

    return redirect("/empleados")

# ================== EDITAR ==================
@app.route("/empleados", methods=["GET", "POST"])
def empleados_view():

    if request.method == "POST":
        conn = get_db()
        cur = conn.cursor()

        empleados = obtener_empleados()

        for emp in empleados:
            zona = request.form.get(f"zona_{emp['id']}")
            puesto = request.form.get(f"puesto_{emp['id']}")
            nombre = request.form.get(f"nombre_{emp['id']}")
            salario = to_float(request.form.get(f"salario_{emp['id']}"))
            tarifa = to_float(request.form.get(f"tarifa_{emp['id']}"))

            cur.execute("""
                UPDATE empleados
                SET zona=%s, puesto=%s, nombre=%s, salario=%s, tarifa=%s
                WHERE id=%s
            """, (zona, puesto, nombre, salario, tarifa, emp["id"]))

        conn.commit()
        cur.close()
        conn.close()

        return redirect("/empleados")

    return render_template("empleados.html", empleados=obtener_empleados())

# ================== HISTORIAL ==================
@app.route("/historial")
def historial():
    carpeta = "nominas"
    if not os.path.exists(carpeta):
        os.makedirs(carpeta)

    archivos = os.listdir(carpeta)
    archivos.sort(reverse=True)

    return render_template("historial.html", archivos=archivos)

@app.route("/descargar/<nombre>")
def descargar(nombre):
    ruta = os.path.join("nominas", nombre)
    return send_file(ruta, as_attachment=True)

@app.route("/eliminar_nomina/<nombre>")
def eliminar_nomina(nombre):
    ruta = os.path.join("nominas", nombre)
    if os.path.exists(ruta):
        os.remove(ruta)
    return redirect("/historial")

# ================== GENERAR EXCEL ==================
@app.route("/generar", methods=["POST"])
def generar():
    try:
        empleados = obtener_empleados()

        if not os.path.exists("nominas"):
            os.makedirs("nominas")

        wb = Workbook()
        ws = wb.active

        # LOGO
        try:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            logo_path = os.path.join(base_dir, "static", "img", "pasi.png")

            if os.path.exists(logo_path):
                img = Image(logo_path)
                img.width = 150
                img.height = 60
                ws.add_image(img, "A1")
                ws.row_dimensions[1].height = 50
        except:
            pass

        # TITULO
        ws.merge_cells("B2:J2")
        ws["B2"] = "PASI - NÓMINA SEMANAL"
        ws["B2"].font = Font(size=14, bold=True)
        ws["B2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("B3:J3")
        ws["B3"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}"
        ws["B3"].alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A6"

        headers = ["Zona","Puesto","Nombre","Base","Tarifa","Horas","Pago T.E","Otros","Descuentos","Total"]

        ws.append([])
        ws.append([])
        ws.append(headers)

        fill = PatternFill(start_color="1F4E78", fill_type="solid")

        for cell in ws[5]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        total_general = 0

        for emp in empleados:
            base = emp["salario"]
            tarifa = emp["tarifa"]

            horas = to_float(request.form.get(f"horas_{emp['id']}"))
            otros = to_float(request.form.get(f"otros_{emp['id']}"))
            desc = to_float(request.form.get(f"desc_{emp['id']}"))

            pago_te = horas * tarifa
            final = base + pago_te + otros - desc

            total_general += final

            ws.append([
                emp["zona"],
                emp["puesto"],
                emp["nombre"],
                base,
                tarifa,
                horas,
                pago_te,
                otros,
                desc,
                final
            ])

        formato = '"$"* #,##0.00_);("$"* (#,##0.00)'

        for row in ws.iter_rows(min_row=6):
            for i in [3,6,7,8,9]:
                row[i].number_format = formato

        ws.append([])
        ws.append(["","","","","","","","","TOTAL", total_general])
        ws[f"J{ws.max_row}"].font = Font(bold=True)

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_len + 2

        archivo = f"nominas/NOMINA_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(archivo)

        return send_file(archivo, as_attachment=True)

    except Exception as e:
        return f"❌ Error: {e}"

# ================== RUN ==================
if __name__ == "__main__":
    app.run(debug=True)